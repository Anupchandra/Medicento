import openpyxl
import re

reader = openpyxl.load_workbook("AIOCD LIST.xlsx")
sheet = reader.active
print(sheet.max_row)
arr = list()
li = ["TAB","CAP","SUS","SUSP","VIAL","VAIL","CAPS","SYP","1VIAL","1DOSE","#","+","O/SUSP","DROPS","TUBE","BOTTLE","N/SPRAY","N/S","SACHET","D/SYP","POWDER","ADHESIVE"]
punc = ["\.","\,","\;"]
for i in range(1,sheet.max_row+1):
    try:
        cell_ = sheet.cell(i,3)
        arr = cell_.value.split()
        if arr[-1].find("S"):
            for item in punc:
                arr[-1] = re.sub(item,"",arr[-1])
        if arr[-1].find("T"):
            arr[-1] = re.sub("T", "S", arr[-1])
        if arr[-1].find("C"):
            arr[-1] = re.sub("C", "S", arr[-1])
        if "OINT" in arr:
            for n, i in enumerate(arr):
                if i == "OINT":
                    arr[n] = "CREAM"
        for item in li:
            if item in arr:
                arr.remove(item)
        print(arr)

    except:
        continue

